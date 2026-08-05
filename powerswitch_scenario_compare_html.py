from __future__ import annotations

import argparse
import csv
import html
import json
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

OUTPUT_COLUMNS = [
    "Execution Date",
    "Network Code",
    "Subnetwork",
    "Region",
    "Area",
    "People",
    "Water heating",
    "General heating",
    "Usage kWh",
    "Retailer",
    "Plan Name",
    "Tariffs and discounts",
    "Plan attributes",
    "Estimated yearly cost",
    "Prices last changed",
    "Result URL",
]

NO_PLANS_LABEL = "No Plans"

TARGET_SCENARIOS = [
    {"network": "VECT", "area": "Auckland Central / Manukau", "people": "1-2 people", "water": "Electric hot water cylinder", "heating": "Heat pump"},
    {"network": "VECT", "area": "Auckland Central / Manukau", "people": "5+ people", "water": "Electric hot water cylinder", "heating": "Heat pump"},
    {"network": "UNET", "area": "Auckland North / West", "people": "1-2 people", "water": "Electric hot water cylinder", "heating": "Heat pump"},
    {"network": "UNET", "area": "Auckland North / West", "people": "5+ people", "water": "Electric hot water cylinder", "heating": "Heat pump"},
    {"network": "CKHK", "area": "Wellington City", "people": "1-2 people", "water": "Electric hot water cylinder", "heating": "Heat pump"},
    {"network": "CKHK", "area": "Wellington City", "people": "5+ people", "water": "Electric hot water cylinder", "heating": "Heat pump"},
    {"network": "ORON", "area": "Christchurch", "people": "1-2 people", "water": "Electric hot water cylinder", "heating": "Heat pump"},
    {"network": "ORON", "area": "Christchurch", "people": "5+ people", "water": "Electric hot water cylinder", "heating": "Heat pump"},
    {"network": "POCO", "area": "Tauranga", "people": "1-2 people", "water": "Electric hot water cylinder", "heating": "Heat pump"},
    {"network": "POCO", "area": "Tauranga", "people": "5+ people", "water": "Electric hot water cylinder", "heating": "Heat pump"},
    {"network": "WAIK", "area": "Hamilton", "people": "1-2 people", "water": "Electric hot water cylinder", "heating": "Heat pump"},
    {"network": "WAIK", "area": "Hamilton", "people": "5+ people", "water": "Electric hot water cylinder", "heating": "Heat pump"},
    {"network": "HAWK", "area": "Hawke's Bay", "people": "1-2 people", "water": "Electric hot water cylinder", "heating": "Heat pump"},
    {"network": "HAWK", "area": "Hawke's Bay", "people": "5+ people", "water": "Electric hot water cylinder", "heating": "Heat pump"},
    {"network": "DUNE", "area": "Dunedin (15 kVA capacity)", "people": "1-2 people", "water": "Electric hot water cylinder", "heating": "Heat pump"},
    {"network": "DUNE", "area": "Dunedin (15 kVA capacity)", "people": "5+ people", "water": "Electric hot water cylinder", "heating": "Heat pump"},
    {"network": "POCO", "area": "Palmerston North", "people": "1-2 people", "water": "Electric hot water cylinder", "heating": "Heat pump"},
    {"network": "POCO", "area": "Palmerston North", "people": "5+ people", "water": "Electric hot water cylinder", "heating": "Heat pump"},
]


@dataclass
class DataFileOption:
    id: str
    label: str
    file_path: Path
    date_key: int
    mtime: float


def clean_string(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_text(value: str) -> str:
    text = clean_string(value).lower()
    text = text.replace("&", " and ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def parse_currency(value: str) -> float | None:
    text = clean_string(value)
    if not text:
        return None
    match = re.search(r"([0-9][0-9,]*(?:\.[0-9]+)?)", text.replace("$", ""))
    if not match:
        return None
    return float(match.group(1).replace(",", ""))


def format_currency(value: float | None) -> str:
    if value is None:
        return ""
    return f"${value:,.2f}"


def is_offer_row(row: dict[str, str]) -> bool:
    plan_name = clean_string(row.get("Plan Name", ""))
    return bool(plan_name) and plan_name != NO_PLANS_LABEL


def scenario_id(scenario: dict[str, str]) -> str:
    return "|".join(
        [
            normalize_text(scenario["network"]),
            normalize_text(scenario["area"]),
            normalize_text(scenario["people"]),
            normalize_text(scenario["water"]),
            normalize_text(scenario["heating"]),
        ]
    )


def row_matches_scenario(row: dict[str, str], scenario: dict[str, str]) -> bool:
    return (
        normalize_text(row.get("Network Code", "")) == normalize_text(scenario["network"])
        and normalize_text(row.get("Area", "")) == normalize_text(scenario["area"])
        and normalize_text(row.get("People", "")) == normalize_text(scenario["people"])
        and normalize_text(row.get("Water heating", "")) == normalize_text(scenario["water"])
        and normalize_text(row.get("General heating", "")) == normalize_text(scenario["heating"])
    )


def extract_date_key(path: Path) -> int:
    # Prefer YYYYMMDD prefix in filename. Fallback to 0 if unavailable.
    match = re.match(r"^(\d{8})", path.stem)
    if not match:
        return 0
    return int(match.group(1))


def is_cumulative_output_file(path: Path) -> bool:
    return path.stem.lower() == "allpowerswitchout"


def load_rows(file_path: Path) -> list[dict[str, str]]:
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        with file_path.open("r", encoding="utf-8", newline="") as f:
            reader = csv.DictReader(f)
            return [{col: clean_string(row.get(col, "")) for col in OUTPUT_COLUMNS} for row in reader]

    if suffix not in {".xlsx", ".xlsm"}:
        raise ValueError(f"Unsupported file type: {file_path}")

    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    if ws.max_row < 1:
        return []

    headers = [clean_string(ws.cell(row=1, column=col_idx).value) for col_idx in range(1, ws.max_column + 1)]
    col_index = {header: idx for idx, header in enumerate(headers)}
    missing = [col for col in OUTPUT_COLUMNS if col not in col_index]
    if missing:
        raise ValueError(f"File {file_path} is missing required columns: {missing}")

    rows: list[dict[str, str]] = []
    for row_idx in range(2, ws.max_row + 1):
        row_dict: dict[str, str] = {}
        for col in OUTPUT_COLUMNS:
            col_idx = col_index[col] + 1
            row_dict[col] = clean_string(ws.cell(row=row_idx, column=col_idx).value)
        if any(row_dict.values()):
            rows.append(row_dict)

    return rows


def discover_data_files(output_folder: Path) -> list[DataFileOption]:
  if not output_folder.exists() or not output_folder.is_dir():
    raise FileNotFoundError(f"Output folder not found: {output_folder}")

  preferred_by_stem: dict[str, Path] = {}
  for path in output_folder.iterdir():
    if not path.is_file():
      continue
    if path.suffix.lower() not in {".xlsx", ".csv"}:
      continue

    stem_lower = path.stem.lower()
    if "powerswitch output" not in stem_lower and not is_cumulative_output_file(path):
      continue

    stem_key = path.stem.lower()
    existing = preferred_by_stem.get(stem_key)
    if existing is None:
      preferred_by_stem[stem_key] = path
      continue

    # Prefer xlsx over csv when both exist for the same scrape period.
    if existing.suffix.lower() == ".csv" and path.suffix.lower() == ".xlsx":
      preferred_by_stem[stem_key] = path

  ranked: list[tuple[int, int, float, Path]] = []
  for path in preferred_by_stem.values():
    ranked.append(
      (
        1 if is_cumulative_output_file(path) else 0,
        extract_date_key(path),
        path.stat().st_mtime,
        path,
      )
    )

  ranked.sort(key=lambda item: (item[0], item[1], item[2], item[3].name.lower()), reverse=True)

  options: list[DataFileOption] = []
  for idx, (_, date_key, mtime, path) in enumerate(ranked, start=1):
    options.append(
      DataFileOption(
        id=f"f{idx}",
        label=path.stem,
        file_path=path,
        date_key=date_key,
        mtime=mtime,
      )
    )

  if not options:
    raise ValueError(f"No scrape output files found in folder: {output_folder}")

  return options


def build_dataset_for_rows(rows: list[dict[str, str]]) -> dict[str, dict[str, Any]]:
    scenarios_map: dict[str, dict[str, Any]] = {}

    for scenario in TARGET_SCENARIOS:
        sid = scenario_id(scenario)
        scenario_rows = [row for row in rows if row_matches_scenario(row, scenario)]
        offers = [row for row in scenario_rows if is_offer_row(row)]

        region = ""
        for row in scenario_rows:
            region = clean_string(row.get("Region", ""))
            if region:
                break

        offer_payload: list[dict[str, Any]] = []
        for row in offers:
            cost = parse_currency(row.get("Estimated yearly cost", ""))
            offer_payload.append(
                {
                    "network": clean_string(row.get("Network Code", "")),
                    "region": clean_string(row.get("Region", "")),
                    "area": clean_string(row.get("Area", "")),
                    "people": clean_string(row.get("People", "")),
                    "water_heating": clean_string(row.get("Water heating", "")),
                    "general_heating": clean_string(row.get("General heating", "")),
                    "usage_kwh": clean_string(row.get("Usage kWh", "")),
                    "retailer": clean_string(row.get("Retailer", "")),
                    "plan_name": clean_string(row.get("Plan Name", "")),
                    "estimated_cost": cost,
                    "estimated_cost_text": format_currency(cost),
                }
            )

        offer_payload.sort(
            key=lambda offer: (
                offer["estimated_cost"] is None,
                offer["estimated_cost"] if offer["estimated_cost"] is not None else 0.0,
                normalize_text(offer["retailer"]),
                normalize_text(offer["plan_name"]),
            )
        )

        has_no_plans = any(clean_string(row.get("Plan Name", "")) == NO_PLANS_LABEL for row in scenario_rows)

        scenarios_map[sid] = {
            "region": region,
            "offers": offer_payload,
            "has_rows": bool(scenario_rows),
            "has_no_plans": has_no_plans,
        }

    return scenarios_map


def build_dashboard_data(file_options: list[DataFileOption]) -> dict[str, Any]:
    datasets: dict[str, dict[str, Any]] = {}

    for option in file_options:
        rows = load_rows(option.file_path)
        datasets[option.id] = {
            "label": option.label,
            "file": str(option.file_path),
            "scenarios": build_dataset_for_rows(rows),
        }

    scenarios_meta = [
        {
            "id": scenario_id(scenario),
            "network": scenario["network"],
            "area": scenario["area"],
            "people": scenario["people"],
            "water": scenario["water"],
            "heating": scenario["heating"],
            "title": f"{scenario['network']} | {scenario['area']} | {scenario['people']} | {scenario['water']} | {scenario['heating']}",
        }
        for scenario in TARGET_SCENARIOS
    ]

    default_current_option = next(
        (option for option in file_options if is_cumulative_output_file(option.file_path)),
        file_options[0],
    )
    default_current = default_current_option.id
    default_compare = next((option.id for option in file_options if option.id != default_current), "")

    return {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "files": [
            {
                "id": option.id,
                "label": option.label,
                "file": str(option.file_path),
            }
            for option in file_options
        ],
        "default_current": default_current,
        "default_compare": default_compare,
        "scenarios": scenarios_meta,
        "datasets": datasets,
    }


def write_dashboard_html(output_file: Path, dashboard: dict[str, Any], output_folder: Path) -> None:
    payload = json.dumps(dashboard, ensure_ascii=False)

    html_content = f"""<!DOCTYPE html>
<html lang=\"en\">
<head>
  <meta charset=\"utf-8\" />
  <meta name=\"viewport\" content=\"width=device-width, initial-scale=1\" />
  <title>PowerSwitch Scenario Plan Comparison</title>
  <style>
    :root {{
      --card: #ffffff;
      --ink: #122433;
      --muted: #607181;
      --line: #d4dee7;
      --accent-soft: #e8f4f6;
      --ok: #0f8f4e;
      --warn: #b06b00;
    }}
    body {{
      margin: 0;
      color: var(--ink);
      font-family: "Segoe UI", "Helvetica Neue", sans-serif;
      background: linear-gradient(180deg, #e9f0f5 0%, #f8fbfd 100%);
    }}
    .wrap {{ max-width: 1760px; margin: 0 auto; padding: 20px; }}
    .hero {{
      background: var(--card);
      border: 1px solid var(--line);
      border-radius: 14px;
      padding: 16px 18px;
      margin-bottom: 14px;
    }}
    h1 {{ margin: 0 0 10px 0; font-size: 28px; }}
    .meta {{ color: var(--muted); font-size: 14px; display: grid; gap: 5px; }}
    .controls {{
      margin-top: 12px;
      display: flex;
      flex-wrap: wrap;
      gap: 10px;
      align-items: flex-start;
    }}
    label {{ font-weight: 600; }}
    select, input {{
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 7px 10px;
      font-size: 14px;
      background: #fff;
      color: var(--ink);
    }}
    #retailerSelect {{ min-width: 250px; min-height: 116px; }}
    .retailer-group {{ display: flex; gap: 8px; align-items: flex-start; }}
    .retailer-buttons {{ display: flex; flex-direction: column; gap: 6px; }}
    .retailer-buttons button {{
      border: 1px solid var(--line);
      border-radius: 8px;
      background: #fff;
      color: var(--ink);
      padding: 6px 10px;
      font-size: 13px;
      cursor: pointer;
    }}
    .retailer-buttons button:hover {{ background: #f4f8fb; }}
    .section {{
      background: var(--card);
      border: 1px solid var(--line);
      border-radius: 14px;
      padding: 12px;
      margin-bottom: 12px;
    }}
    .section h2 {{ margin: 0 0 8px 0; font-size: 17px; }}
    .table-wrap {{
      overflow: auto;
      border: 1px solid var(--line);
      border-radius: 10px;
      max-height: 62vh;
      background: #fff;
    }}
    table {{ border-collapse: collapse; min-width: 1660px; width: 100%; }}
    th, td {{
      border-bottom: 1px solid #edf2f7;
      padding: 7px 9px;
      white-space: nowrap;
      font-size: 12.5px;
      text-align: left;
    }}
    th {{
      position: sticky;
      top: 0;
      z-index: 1;
      background: var(--accent-soft);
      border-bottom: 1px solid #c7dae0;
      font-weight: 700;
    }}
    tbody tr:nth-child(even) {{ background: #fbfdff; }}
    td.currency, td.delta {{ text-align: right; }}
    .match-ok {{ color: var(--ok); font-weight: 600; }}
    .match-warn {{ color: var(--warn); font-weight: 600; }}
    .small-note {{ color: var(--muted); font-size: 12px; margin-top: 6px; }}
  </style>
</head>
<body>
  <div class=\"wrap\">
    <section class=\"hero\">
      <h1>PowerSwitch Scenario Plan Comparison</h1>
      <div class=\"meta\">
        <div><strong>Output folder:</strong> {html.escape(str(output_folder))}</div>
        <div><strong>Generated:</strong> {html.escape(dashboard['generated_at'])}</div>
      </div>
      <div class=\"controls\">
        <label for=\"currentFileSelect\">Current file:</label>
        <select id=\"currentFileSelect\"></select>

        <label for=\"compareFileSelect\">Compare file:</label>
        <select id=\"compareFileSelect\"></select>

        <label for=\"retailerSelect\">Retailers:</label>
        <div class=\"retailer-group\">
          <select id=\"retailerSelect\" multiple></select>
          <div class=\"retailer-buttons\">
            <button id=\"retailerAll\" type=\"button\">All</button>
            <button id=\"retailerNone\" type=\"button\">None</button>
          </div>
        </div>

        <label for=\"globalFilter\">Filter text:</label>
        <input id=\"globalFilter\" type=\"text\" placeholder=\"Retailer, plan, area...\" />
      </div>
      <div class=\"small-note\">Fuzzy matching is used for plan names. If no similar plan is found, match status shows No similar plan.</div>
    </section>

    <div id=\"tablesRoot\"></div>
  </div>

  <script>
    const DATA = {payload};

    function money(value) {{
      if (value === null || value === undefined || value === "") return "";
      const num = Number(value);
      if (Number.isNaN(num)) return "";
      return "$" + num.toLocaleString(undefined, {{ minimumFractionDigits: 2, maximumFractionDigits: 2 }});
    }}

    function normalizeText(value) {{
      return String(value || "")
        .toLowerCase()
        .replace(/&/g, " and ")
        .replace(/[^a-z0-9]+/g, " ")
        .replace(/\\s+/g, " ")
        .trim();
    }}

    function normalizePlanName(planName) {{
      const stop = new Set(["plan", "power", "electricity", "offer", "bundle", "fixed", "variable", "price"]);
      const norm = normalizeText(planName);
      const parts = norm.split(" ").filter(Boolean).filter((p) => !stop.has(p));
      return (parts.length ? parts : norm.split(" ").filter(Boolean)).join(" ");
    }}

    function bigrams(text) {{
      if (!text) return [];
      if (text.length === 1) return [text];
      const out = [];
      for (let i = 0; i < text.length - 1; i += 1) out.push(text.slice(i, i + 2));
      return out;
    }}

    function diceCoefficient(a, b) {{
      if (!a || !b) return 0;
      const aa = bigrams(a);
      const bb = bigrams(b);
      const counts = new Map();
      for (const token of aa) counts.set(token, (counts.get(token) || 0) + 1);
      let overlap = 0;
      for (const token of bb) {{
        const c = counts.get(token) || 0;
        if (c > 0) {{
          overlap += 1;
          counts.set(token, c - 1);
        }}
      }}
      return (2 * overlap) / (aa.length + bb.length);
    }}

    function nameSimilarity(a, b) {{
      const na = normalizePlanName(a);
      const nb = normalizePlanName(b);
      if (!na || !nb) return 0;

      const ratio = diceCoefficient(na, nb);
      const sa = new Set(na.split(" "));
      const sb = new Set(nb.split(" "));
      let inter = 0;
      for (const token of sa) if (sb.has(token)) inter += 1;
      const union = new Set([...sa, ...sb]).size || 1;
      const overlap = inter / union;

      return 0.75 * ratio + 0.25 * overlap;
    }}

    function matchCurrentToBaseline(currentOffers, baselineOffers) {{
      if (!baselineOffers || baselineOffers.length === 0) {{
        return currentOffers.map(() => ({{ old_cost: null, change_cost: null, match_status: "No baseline scenario" }}));
      }}

      const pairs = [];
      for (let ci = 0; ci < currentOffers.length; ci += 1) {{
        const cur = currentOffers[ci];
        const curRetailer = normalizeText(cur.retailer);
        for (let bi = 0; bi < baselineOffers.length; bi += 1) {{
          const old = baselineOffers[bi];
          const oldRetailer = normalizeText(old.retailer);
          const sameRetailer = Boolean(curRetailer) && curRetailer === oldRetailer;

          let score = nameSimilarity(cur.plan_name, old.plan_name);
          score += sameRetailer ? 0.08 : -0.08;
          score = Math.max(0, Math.min(1, score));

          const threshold = sameRetailer ? 0.62 : 0.86;
          if (score >= threshold) pairs.push({{ score, sameRetailer, ci, bi }});
        }}
      }}

      pairs.sort((a, b) => {{
        if (b.score !== a.score) return b.score - a.score;
        if (b.sameRetailer !== a.sameRetailer) return Number(b.sameRetailer) - Number(a.sameRetailer);
        return 0;
      }});

      const usedCurrent = new Set();
      const usedBaseline = new Set();
      const assignment = new Map();

      for (const pair of pairs) {{
        if (usedCurrent.has(pair.ci) || usedBaseline.has(pair.bi)) continue;
        usedCurrent.add(pair.ci);
        usedBaseline.add(pair.bi);
        assignment.set(pair.ci, pair);
      }}

      return currentOffers.map((cur, ci) => {{
        const picked = assignment.get(ci);
        if (!picked) {{
          return {{ old_cost: null, change_cost: null, match_status: "No similar plan" }};
        }}

        const old = baselineOffers[picked.bi];
        const oldCost = old.estimated_cost;
        const curCost = cur.estimated_cost;
        const change = oldCost === null || oldCost === undefined || curCost === null || curCost === undefined
          ? null
          : curCost - oldCost;

        return {{
          old_cost: oldCost,
          change_cost: change,
          match_status: `Matched (${{picked.score.toFixed(2)}})`,
        }};
      }});
    }}

    function optionById(id) {{
      return DATA.files.find((f) => f.id === id);
    }}

    function getCurrentFileId() {{
      return document.getElementById("currentFileSelect").value;
    }}

    function getCompareFileId() {{
      return document.getElementById("compareFileSelect").value;
    }}

    function getScenarioData(fileId, scenarioId) {{
      if (!fileId) return {{ region: "", offers: [], has_rows: false, has_no_plans: false }};
      const ds = DATA.datasets[fileId];
      if (!ds) return {{ region: "", offers: [], has_rows: false, has_no_plans: false }};
      return ds.scenarios[scenarioId] || {{ region: "", offers: [], has_rows: false, has_no_plans: false }};
    }}

    function buildFileSelectors() {{
      const currentSelect = document.getElementById("currentFileSelect");
      const compareSelect = document.getElementById("compareFileSelect");

      currentSelect.innerHTML = "";
      compareSelect.innerHTML = "";

      for (const file of DATA.files) {{
        const option = document.createElement("option");
        option.value = file.id;
        option.textContent = file.label;
        currentSelect.appendChild(option);
      }}

      const none = document.createElement("option");
      none.value = "";
      none.textContent = "No comparison";
      compareSelect.appendChild(none);

      for (const file of DATA.files) {{
        const option = document.createElement("option");
        option.value = file.id;
        option.textContent = file.label;
        compareSelect.appendChild(option);
      }}

      currentSelect.value = DATA.default_current || (DATA.files[0] ? DATA.files[0].id : "");
      compareSelect.value = DATA.default_compare || "";
    }}

    function collectRetailersForCurrentFile() {{
      const currentId = getCurrentFileId();
      const set = new Set();
      for (const scenario of DATA.scenarios) {{
        const data = getScenarioData(currentId, scenario.id);
        for (const offer of data.offers || []) {{
          const retailer = (offer.retailer || "").trim();
          if (retailer) set.add(retailer);
        }}
      }}
      return Array.from(set).sort((a, b) => a.localeCompare(b));
    }}

    function buildRetailerSelect(preserveSelection) {{
      const select = document.getElementById("retailerSelect");
      const previous = new Set(
        Array.from(select.options)
          .filter((opt) => opt.selected)
          .map((opt) => opt.value)
      );

      select.innerHTML = "";
      const retailers = collectRetailersForCurrentFile();
      for (const retailer of retailers) {{
        const opt = document.createElement("option");
        opt.value = retailer;
        opt.textContent = retailer;
        opt.selected = preserveSelection ? previous.has(retailer) : true;
        if (!preserveSelection) opt.selected = true;
        select.appendChild(opt);
      }}

      if (preserveSelection && previous.size === 0) {{
        for (const opt of select.options) opt.selected = true;
      }}
      if (!preserveSelection) {{
        for (const opt of select.options) opt.selected = true;
      }}
    }}

    function selectedRetailers() {{
      const select = document.getElementById("retailerSelect");
      const out = new Set();
      for (const opt of select.options) if (opt.selected) out.add(opt.value);
      return out;
    }}

    function addCell(row, text, className) {{
      const td = document.createElement("td");
      td.textContent = text || "";
      if (className) td.className = className;
      row.appendChild(td);
    }}

    function headerColumns() {{
      return [
        "Network",
        "Region",
        "Area",
        "People",
        "Water Heating",
        "General Heating",
        "Usage kWh",
        "Retailer",
        "Plan Name",
        "Estimated Yearly Cost",
        "Variance to Cheapest",
        "Old Cost",
        "Change in Cost",
        "Match Status",
      ];
    }}

    function renderTables() {{
      const currentId = getCurrentFileId();
      const compareId = getCompareFileId();
      const root = document.getElementById("tablesRoot");
      root.innerHTML = "";

      for (const scenario of DATA.scenarios) {{
        const currentData = getScenarioData(currentId, scenario.id);
        const baselineData = getScenarioData(compareId, scenario.id);

        const currentOffers = [...(currentData.offers || [])].sort((a, b) => {{
          const ac = a.estimated_cost;
          const bc = b.estimated_cost;
          if (ac == null && bc == null) return 0;
          if (ac == null) return 1;
          if (bc == null) return -1;
          return ac - bc;
        }});

        const numeric = currentOffers.map((r) => r.estimated_cost).filter((c) => c != null);
        const cheapest = numeric.length ? Math.min(...numeric) : null;

        const compareMap = compareId ? matchCurrentToBaseline(currentOffers, baselineData.offers || []) : [];

        const section = document.createElement("section");
        section.className = "section";

        const h2 = document.createElement("h2");
        h2.textContent = scenario.title;
        section.appendChild(h2);

        const wrap = document.createElement("div");
        wrap.className = "table-wrap";
        const table = document.createElement("table");

        const thead = document.createElement("thead");
        const trHead = document.createElement("tr");
        for (const title of headerColumns()) {{
          const th = document.createElement("th");
          th.textContent = title;
          trHead.appendChild(th);
        }}
        thead.appendChild(trHead);
        table.appendChild(thead);

        const tbody = document.createElement("tbody");
        if (!currentOffers.length) {{
          const tr = document.createElement("tr");
          tr.dataset.role = "plan-row";
          tr.dataset.retailer = "";

          let message = "Scenario not present in selected current file";
          if (currentData.has_rows && currentData.has_no_plans) message = "No plans in selected current file";

          const region = currentData.region || baselineData.region || "";

          addCell(tr, scenario.network);
          addCell(tr, region);
          addCell(tr, scenario.area);
          addCell(tr, scenario.people);
          addCell(tr, scenario.water);
          addCell(tr, scenario.heating);
          addCell(tr, "");
          addCell(tr, "");
          addCell(tr, message);
          addCell(tr, "", "currency");
          addCell(tr, "", "currency");
          addCell(tr, "", "currency");
          addCell(tr, "", "delta");
          addCell(tr, compareId ? "No latest plan" : "");

          tr.dataset.search = normalizeText(tr.innerText || "");
          tbody.appendChild(tr);
        }} else {{
          for (let i = 0; i < currentOffers.length; i += 1) {{
            const row = currentOffers[i];
            const compare = compareId ? compareMap[i] : null;
            const variance = row.estimated_cost == null || cheapest == null ? null : row.estimated_cost - cheapest;

            const tr = document.createElement("tr");
            tr.dataset.role = "plan-row";
            tr.dataset.retailer = row.retailer || "";

            addCell(tr, row.network || scenario.network);
            addCell(tr, row.region || currentData.region || baselineData.region || "");
            addCell(tr, row.area || scenario.area);
            addCell(tr, row.people || scenario.people);
            addCell(tr, row.water_heating || scenario.water);
            addCell(tr, row.general_heating || scenario.heating);
            addCell(tr, row.usage_kwh || "");
            addCell(tr, row.retailer || "");
            addCell(tr, row.plan_name || "");
            addCell(tr, money(row.estimated_cost), "currency");
            addCell(tr, money(variance), "currency");

            const oldCost = compare ? money(compare.old_cost) : "";
            const changeCost = compare ? money(compare.change_cost) : "";
            const status = compare ? (compare.match_status || "") : "";

            addCell(tr, oldCost, "currency");
            addCell(tr, changeCost, "delta");

            const statusCell = document.createElement("td");
            statusCell.textContent = status;
            if (status.startsWith("Matched")) statusCell.className = "match-ok";
            else if (status) statusCell.className = "match-warn";
            tr.appendChild(statusCell);

            tr.dataset.search = normalizeText(tr.innerText || "");
            tbody.appendChild(tr);
          }}
        }}

        table.appendChild(tbody);
        wrap.appendChild(table);
        section.appendChild(wrap);

        const note = document.createElement("div");
        note.className = "small-note";
        note.dataset.role = "section-count";
        note.textContent = `Rows: ${{tbody.querySelectorAll('tr').length.toLocaleString()}}`;
        section.appendChild(note);

        root.appendChild(section);
      }}

      applyRowFilters();
    }}

    function applyRowFilters() {{
      const selected = selectedRetailers();
      const query = normalizeText(document.getElementById("globalFilter").value || "");

      for (const section of document.querySelectorAll("section.section")) {{
        const rows = section.querySelectorAll("tbody tr[data-role='plan-row']");
        let visible = 0;
        for (const tr of rows) {{
          const retailer = (tr.dataset.retailer || "").trim();
          const retailerPass = retailer ? selected.has(retailer) : true;
          const text = tr.dataset.search || normalizeText(tr.innerText || "");
          const textPass = !query || text.includes(query);
          const show = retailerPass && textPass;
          tr.style.display = show ? "" : "none";
          if (show) visible += 1;
        }}

        const note = section.querySelector("[data-role='section-count']");
        if (note) {{
          note.textContent = `Visible rows: ${{visible.toLocaleString()}} / ${{rows.length.toLocaleString()}}`;
        }}
      }}
    }}

    function selectAllRetailers(state) {{
      const select = document.getElementById("retailerSelect");
      for (const opt of select.options) opt.selected = state;
      applyRowFilters();
    }}

    function init() {{
      buildFileSelectors();
      buildRetailerSelect(false);
      renderTables();

      document.getElementById("currentFileSelect").addEventListener("change", () => {{
        buildRetailerSelect(false);
        renderTables();
      }});

      document.getElementById("compareFileSelect").addEventListener("change", () => {{
        renderTables();
      }});

      document.getElementById("retailerSelect").addEventListener("change", applyRowFilters);
      document.getElementById("globalFilter").addEventListener("input", applyRowFilters);
      document.getElementById("retailerAll").addEventListener("click", () => selectAllRetailers(true));
      document.getElementById("retailerNone").addEventListener("click", () => selectAllRetailers(false));
    }}

    init();
  </script>
</body>
</html>
"""

    output_file.write_text(html_content, encoding="utf-8")


def build_default_output_html(file_options: list[DataFileOption]) -> Path:
    latest = next((option for option in file_options if option.date_key), file_options[0])
    match = re.match(r"^(\d{8})\b", latest.label)
    if match:
        date_text = match.group(1)
    elif latest.date_key:
        date_text = f"{latest.date_key:08d}"
    else:
        date_text = datetime.now().strftime("%Y%m%d")
    return Path(f"{date_text} Powerswitch Scenario Comparison.html")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build scenario comparison HTML from all scrape files in the Output folder."
    )
    parser.add_argument(
        "--output-folder",
        default="Output",
        help="Folder containing scrape output files to compare.",
    )
    parser.add_argument(
        "--output-html",
        default=None,
        help="Output HTML file path. Default uses latest scrape date from the output folder.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    output_folder = Path(args.output_folder)

    file_options = discover_data_files(output_folder)
    output_html = Path(args.output_html) if args.output_html else build_default_output_html(file_options)
    dashboard = build_dashboard_data(file_options)
    write_dashboard_html(output_html, dashboard=dashboard, output_folder=output_folder)

    current_option = next((option for option in file_options if option.id == dashboard["default_current"]), None)
    compare_option = next((option for option in file_options if option.id == dashboard["default_compare"]), None)

    print("Scenario comparison dashboard created")
    print(f"Output folder: {output_folder}")
    print(f"Files discovered: {len(file_options)}")
    if current_option:
        print(f"Default current: {current_option.label}")
    if compare_option:
        print(f"Default compare: {compare_option.label}")
    print(f"Output HTML: {output_html}")


if __name__ == "__main__":
    main()
