from __future__ import annotations

import argparse
from concurrent.futures import ProcessPoolExecutor, as_completed
import json
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from playwright.sync_api import Page, TimeoutError as PlaywrightTimeoutError, sync_playwright

OUTPUT_COLUMNS = [
    "Execution Date",
    "Network Code",
    "Subnetwork",
    "Region",
    "Area",
    "People",
    "Water heating",
    "General heating",
    "Usage kWh",
    "Retailer",
    "Plan Name",
    "Tariffs and discounts",
    "Plan attributes",
    "Estimated yearly cost",
    "Prices last changed",
    "Result URL",
]

DEFAULT_OUTPUT_DIR = Path("Output")
DEFAULT_ALL_OUTPUT_CSV = DEFAULT_OUTPUT_DIR / "ALLPOWERSWITCHOUT.csv"


@dataclass
class Scenario:
    row_num: int
    network_code: str
    subnetwork: str
    region: str
    area: str
    people: str
    water: str
    heating: str


def clean_string(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_row_list(row_list: str) -> list[int]:
    rows: list[int] = []
    for token in [part.strip() for part in row_list.split(",") if part.strip()]:
        if "-" in token:
            start_text, end_text = token.split("-", 1)
            start_num = int(start_text)
            end_num = int(end_text)
            if start_num > end_num:
                raise ValueError(f"Invalid row range '{token}'")
            rows.extend(range(start_num, end_num + 1))
        else:
            rows.append(int(token))

    deduped: list[int] = []
    seen: set[int] = set()
    for row_num in rows:
        if row_num < 1:
            raise ValueError(f"Invalid row number '{row_num}' in --row-list")
        if row_num in seen:
            continue
        seen.add(row_num)
        deduped.append(row_num)

    if not deduped:
        raise ValueError("--row-list did not contain any valid rows")

    return deduped


def resolve_output_path(output_file_arg: str) -> Path:
    output_path = Path(output_file_arg)
    if output_path.is_absolute():
        return output_path
    if output_path.parent != Path("."):
        return output_path
    return DEFAULT_OUTPUT_DIR / output_path.name


def build_no_plans_placeholder_row(
    scenario: Scenario,
    execution_date: str,
    usage: str,
    result_url: str,
    note: str,
) -> dict[str, str]:
    return {
        "Execution Date": execution_date,
        "Network Code": scenario.network_code,
        "Subnetwork": scenario.subnetwork,
        "Region": scenario.region,
        "Area": scenario.area,
        "People": scenario.people,
        "Water heating": scenario.water,
        "General heating": scenario.heating,
        "Usage kWh": usage,
        "Retailer": "",
        "Plan Name": "No Plans",
        "Tariffs and discounts": "not applicable",
        "Plan attributes": note,
        "Estimated yearly cost": "",
        "Prices last changed": "",
        "Result URL": result_url,
    }


def load_scenarios(
    spec_file: Path,
    limit_rows: int,
    start_row: int = 2,
    row_numbers: list[int] | None = None,
) -> list[Scenario]:
    wb = load_workbook(spec_file, data_only=True)
    ws = wb.active

    if row_numbers:
        scenarios: list[Scenario] = []
        for row_idx in row_numbers:
            if row_idx > ws.max_row:
                raise ValueError(f"Row {row_idx} does not exist in spec workbook (max row: {ws.max_row})")

            network_code = clean_string(ws[f"A{row_idx}"].value)
            subnetwork = clean_string(ws[f"B{row_idx}"].value)
            region = clean_string(ws[f"C{row_idx}"].value)
            area = clean_string(ws[f"D{row_idx}"].value)
            people = clean_string(ws[f"E{row_idx}"].value)
            water = clean_string(ws[f"F{row_idx}"].value)
            heating = clean_string(ws[f"G{row_idx}"].value)

            values = [network_code, subnetwork, region, area, people, water, heating]
            if not any(values):
                raise ValueError(f"Row {row_idx} is blank in columns A-G")
            if not all(values):
                raise ValueError(f"Row {row_idx} is missing one or more required values in columns A-G")

            scenarios.append(
                Scenario(
                    row_num=row_idx,
                    network_code=network_code,
                    subnetwork=subnetwork,
                    region=region,
                    area=area,
                    people=people,
                    water=water,
                    heating=heating,
                )
            )

        return scenarios

    scenarios: list[Scenario] = []
    for row_idx in range(start_row, ws.max_row + 1):
        network_code = clean_string(ws[f"A{row_idx}"].value)
        subnetwork = clean_string(ws[f"B{row_idx}"].value)
        region = clean_string(ws[f"C{row_idx}"].value)
        area = clean_string(ws[f"D{row_idx}"].value)
        people = clean_string(ws[f"E{row_idx}"].value)
        water = clean_string(ws[f"F{row_idx}"].value)
        heating = clean_string(ws[f"G{row_idx}"].value)

        values = [network_code, subnetwork, region, area, people, water, heating]
        if not any(values):
            continue
        if not all(values):
            raise ValueError(f"Row {row_idx} is missing one or more required values in columns A-G")

        scenarios.append(
            Scenario(
                row_num=row_idx,
                network_code=network_code,
                subnetwork=subnetwork,
                region=region,
                area=area,
                people=people,
                water=water,
                heating=heating,
            )
        )

        if len(scenarios) >= limit_rows:
            break

    if len(scenarios) < limit_rows:
        raise ValueError(
            f"Requested {limit_rows} rows starting at {start_row}, but only found {len(scenarios)} usable rows"
        )

    return scenarios


def maybe_click(locator, timeout: int = 3000) -> bool:
    try:
        if locator.count() > 0:
            locator.first.click(timeout=timeout)
            return True
    except Exception:
        return False
    return False


def click_main_text_exact(page: Page, text: str, timeout_ms: int) -> None:
    rx = re.compile(rf"^{re.escape(text)}$", re.IGNORECASE)
    loc = page.locator("main").get_by_text(rx).first
    loc.wait_for(timeout=timeout_ms)
    loc.click(timeout=timeout_ms)


def close_info_popup_if_needed(page: Page) -> None:
    close_btn = page.get_by_role("button", name=re.compile(r"Close info popup", re.IGNORECASE))
    for _ in range(2):
        if not maybe_click(close_btn, timeout=1500):
            break


def configure_context_for_speed(context, block_assets: bool) -> None:
    if not block_assets:
        return

    blocked_types = {"image", "font", "media"}

    def route_handler(route) -> None:
        if route.request.resource_type in blocked_types:
            route.abort()
        else:
            route.continue_()

    context.route("**/*", route_handler)


def expand_all_plan_details(page: Page, timeout_ms: int) -> None:
    details = page.get_by_role("button", name=re.compile(r"^Plan details$", re.IGNORECASE))
    count = details.count()
    for idx in range(count):
        btn = details.nth(idx)
        try:
            expanded = (btn.get_attribute("aria-expanded") or "").lower()
            if expanded != "true":
                btn.click(timeout=timeout_ms)
        except Exception:
            # Continue if one plan card fails to expand.
            pass


def run_questionnaire_flow(page: Page, scenario: Scenario, timeout_ms: int) -> None:
    page.goto("https://www.powerswitch.org.nz/", wait_until="domcontentloaded")
    maybe_click(page.get_by_role("button", name=re.compile(r"Close and accept cookie policy", re.IGNORECASE)))

    page.get_by_role("link", name=re.compile(r"^Skip$", re.IGNORECASE)).first.click(timeout=timeout_ms)
    maybe_click(page.get_by_role("button", name=re.compile(r"Close and accept cookie policy", re.IGNORECASE)))

    page.get_by_role("heading", name=re.compile(r"What region do you live in\?", re.IGNORECASE)).wait_for(
        timeout=timeout_ms
    )
    click_main_text_exact(page, scenario.region, timeout_ms)

    area_box = page.get_by_role("combobox").first
    area_box.wait_for(timeout=timeout_ms)
    area_box.fill("")
    area_box.type(scenario.area, delay=20)
    try:
        page.get_by_role("option", name=scenario.area, exact=True).first.click(timeout=timeout_ms)
    except Exception:
        page.get_by_role("option", name=re.compile(re.escape(scenario.area), re.IGNORECASE)).first.click(
            timeout=timeout_ms
        )
    page.get_by_role("button", name=re.compile(r"^Next step$", re.IGNORECASE)).click(timeout=timeout_ms)

    page.get_by_role("heading", name=re.compile(r"Your household", re.IGNORECASE)).wait_for(timeout=timeout_ms)
    page.get_by_role("button", name=re.compile(r"^Skip$", re.IGNORECASE)).click(timeout=timeout_ms)

    page.get_by_role("heading", name=re.compile(r"Do you use gas\?", re.IGNORECASE)).wait_for(timeout=timeout_ms)
    click_main_text_exact(page, "None", timeout_ms)
    page.get_by_role("button", name=re.compile(r"^Next step$", re.IGNORECASE)).click(timeout=timeout_ms)

    page.get_by_role("heading", name=re.compile(r"How many people live in your home\?", re.IGNORECASE)).wait_for(
        timeout=timeout_ms
    )
    click_main_text_exact(page, scenario.people, timeout_ms)
    click_main_text_exact(page, "No", timeout_ms)
    page.get_by_role("button", name=re.compile(r"^Next step$", re.IGNORECASE)).click(timeout=timeout_ms)

    page.get_by_role(
        "heading", name=re.compile(r"Do you use any of these to heat your water\?", re.IGNORECASE)
    ).wait_for(timeout=timeout_ms)
    click_main_text_exact(page, scenario.water, timeout_ms)
    page.get_by_role("button", name=re.compile(r"^Next step$", re.IGNORECASE)).click(timeout=timeout_ms)

    page.get_by_role(
        "heading", name=re.compile(r"Do you use any of these to heat your home\?", re.IGNORECASE)
    ).wait_for(timeout=timeout_ms)
    click_main_text_exact(page, scenario.heating, timeout_ms)
    if "heat pump" in scenario.heating.lower():
        click_main_text_exact(page, "Yes", timeout_ms)
    page.get_by_role("button", name=re.compile(r"^Next step$", re.IGNORECASE)).click(timeout=timeout_ms)

    page.get_by_role("heading", name=re.compile(r"What insulation do you have\?", re.IGNORECASE)).wait_for(
        timeout=timeout_ms
    )
    click_main_text_exact(page, "Underfloor", timeout_ms)
    click_main_text_exact(page, "Wall", timeout_ms)
    click_main_text_exact(page, "Ceiling", timeout_ms)
    page.get_by_role("button", name=re.compile(r"^View results$", re.IGNORECASE)).click(timeout=timeout_ms)

    page.get_by_role("heading", name=re.compile(r"Available plans", re.IGNORECASE)).wait_for(
        timeout=max(timeout_ms, 120_000)
    )


def extract_rows_for_scenario(page: Page, scenario: Scenario, execution_date: str) -> tuple[list[dict[str, str]], str, str]:
    payload = {
        "executionDate": execution_date,
        "scenario": {
            "networkCode": scenario.network_code,
            "subnetwork": scenario.subnetwork,
            "region": scenario.region,
            "area": scenario.area,
            "people": scenario.people,
            "water": scenario.water,
            "heating": scenario.heating,
        },
    }

    result = page.evaluate(
        """
        ({ executionDate, scenario }) => {
          const clean = (str) => (str || '').replace(/\\u00a0/g, ' ').replace(/\\s+/g, ' ').trim();
          const unique = (arr) => {
            const out = [];
            for (const v of arr) {
              if (v && !out.includes(v)) out.push(v);
            }
            return out;
          };

          const usageRow = Array.from(document.querySelectorAll('tr')).find((tr) =>
            /Estimated usage:/i.test(tr.innerText || '')
          );
          let usage = '';
          if (usageRow) {
            const txt = clean(usageRow.innerText || '');
            const m = txt.match(/Estimated usage:\\s*([0-9,]+)/i);
            if (m) usage = m[1];
          }

          const cards = Array.from(document.querySelectorAll('li')).filter((li) => {
            const t = li.innerText || '';
            return li.querySelector('h3') && /Estimated yearly cost/i.test(t);
          });

          const rows = cards
            .map((li) => {
              const fullText = clean(li.innerText || '');

              const logoAlt = clean(li.querySelector('img[alt$="logo"]')?.getAttribute('alt') || '');
              let retailer = '';
              if (logoAlt) retailer = logoAlt.replace(/\\s+logo$/i, '').trim();
              if (!retailer) retailer = clean(li.querySelector('a[href*="/retailers/"]')?.textContent || '');

              const planName = clean(li.querySelector('h3')?.textContent || '');

              let estimated = '';
                            const mCost =
                                fullText.match(/Estimated yearly cost\\s*\\$\\s*([0-9,]+\\*?)/i) ||
                                fullText.match(/\\$\\s*([0-9,]+\\*?)/);
              if (mCost) estimated = '$' + mCost[1].replace(/^\\$/, '');

              const tariffPairs = [];
              for (const tr of Array.from(li.querySelectorAll('table tr'))) {
                const tds = tr.querySelectorAll('td');
                if (tds.length === 2) {
                  const key = clean(tds[0].innerText);
                  const val = clean(tds[1].innerText);
                  if (key && val) tariffPairs.push(`${key}: ${val}`);
                }
              }
              const tariffs = tariffPairs.length ? `${tariffPairs.join(' | ')} | ` : 'not applicable';

              const exclude = /^(switch|add to compare|tariffs and discounts|plan details|show more|show less|how do you calculate power cost and what does the estimate include\\?|compare|select plans to compare|select to compare|cheapest|bundle plan|i)$/i;
              const attrs = unique(
                Array.from(li.querySelectorAll('ul li button'))
                  .map((b) => clean(b.textContent))
                  .map((t) => t.replace(/^i(?=[A-Z])/, ''))
                  .filter(
                    (t) =>
                      t &&
                      !exclude.test(t) &&
                      t.length <= 90 &&
                      !/^\\$[0-9,]+\\*?$/.test(t) &&
                      !/^incl\\. GST$/i.test(t)
                  )
              ).join(' | ');

              let pricesLastChanged = '';
              const mPrice = fullText.match(/Prices last changed\\s*([A-Za-z]+\\s+[0-9]{4})/i);
              if (mPrice) pricesLastChanged = mPrice[1] + ' ';

              return {
                'Execution Date': executionDate,
                'Network Code': scenario.networkCode,
                Subnetwork: scenario.subnetwork,
                Region: scenario.region,
                Area: scenario.area,
                People: scenario.people,
                'Water heating': scenario.water,
                'General heating': scenario.heating,
                'Usage kWh': usage,
                Retailer: retailer,
                'Plan Name': planName,
                'Tariffs and discounts': tariffs,
                'Plan attributes': attrs,
                'Estimated yearly cost': estimated,
                'Prices last changed': pricesLastChanged,
                'Result URL': location.href,
              };
            })
            .filter((row) => row['Plan Name']);

          return {
            usage,
            resultUrl: location.href,
            rowCount: rows.length,
            rows,
          };
        }
        """,
        payload,
    )

    rows: list[dict[str, str]] = result["rows"]
    return rows, result["resultUrl"], result["usage"]


def safe_close(target: Any) -> None:
    try:
        target.close()
    except Exception:
        pass


def open_worker_browser_stack(playwright: Any, show_browser: bool, block_assets: bool):
    browser = playwright.chromium.launch(headless=not show_browser)
    context = browser.new_context()
    configure_context_for_speed(context, block_assets=block_assets)
    page = context.new_page()
    return browser, context, page


def recover_worker_page(
    playwright: Any,
    browser: Any,
    context: Any,
    page: Any,
    show_browser: bool,
    block_assets: bool,
    worker_id: int,
):
    safe_close(page)
    print(
        f"    !! worker {worker_id} recycling browser stack after attempt failure",
        flush=True,
    )
    safe_close(context)
    safe_close(browser)
    return open_worker_browser_stack(playwright, show_browser=show_browser, block_assets=block_assets)


def process_scenario_batch(
    scenarios: list[Scenario],
    timeout_ms: int,
    show_browser: bool,
    execution_date: str,
    block_assets: bool,
    worker_id: int,
    max_retries: int,
    retry_empty_results: bool,
    fail_fast: bool,
) -> tuple[list[dict[str, str]], list[dict[str, str | int]]]:
    all_rows: list[dict[str, str]] = []
    scenario_summaries: list[dict[str, str | int]] = []

    with sync_playwright() as playwright:
        browser, context, page = open_worker_browser_stack(
            playwright,
            show_browser=show_browser,
            block_assets=block_assets,
        )

        for idx, scenario in enumerate(scenarios, start=1):
            print(
                f"[W{worker_id} {idx}/{len(scenarios)}] Row {scenario.row_num}: "
                f"{scenario.region} / {scenario.area} / {scenario.people} / {scenario.water} / {scenario.heating}"
                ,
                flush=True,
            )

            attempt_limit = max(1, max_retries + 1)
            completed = False
            last_error = ""
            last_usage = ""
            last_result_url = ""

            for attempt in range(1, attempt_limit + 1):
                try:
                    run_questionnaire_flow(page, scenario, timeout_ms=timeout_ms)
                    close_info_popup_if_needed(page)
                    expand_all_plan_details(page, timeout_ms=timeout_ms)

                    rows, result_url, usage = extract_rows_for_scenario(page, scenario, execution_date)
                    last_result_url = result_url
                    last_usage = usage
                    if retry_empty_results and len(rows) == 0 and attempt < attempt_limit:
                        raise RuntimeError("No plans captured on results page")

                    output_rows: list[dict[str, str]] = rows
                    if len(rows) == 0:
                        output_rows = [
                            build_no_plans_placeholder_row(
                                scenario=scenario,
                                execution_date=execution_date,
                                usage=usage,
                                result_url=result_url,
                                note="No plans returned by PowerSwitch",
                            )
                        ]

                    for plan_idx, row in enumerate(output_rows):
                        row["__row_num"] = str(scenario.row_num)
                        row["__plan_order"] = str(plan_idx)

                    all_rows.extend(output_rows)
                    scenario_summaries.append(
                        {
                            "row_num": scenario.row_num,
                            "result_url": result_url,
                            "usage": usage,
                            "plan_count": len(rows),
                            "status": "ok" if rows else "empty",
                        }
                    )
                    print(f"    -> {len(rows)} plans captured, usage {usage}, URL {result_url}", flush=True)
                    completed = True
                    break

                except Exception as exc:
                    last_error = f"{type(exc).__name__}: {exc}"
                    print(
                        f"    !! attempt {attempt}/{attempt_limit} failed for row {scenario.row_num}: {last_error}",
                        flush=True,
                    )

                    screenshot_name = (
                        f"powerswitch_scrape_timeout_worker{worker_id}_row{scenario.row_num}_attempt{attempt}.png"
                    )
                    try:
                        page.screenshot(path=screenshot_name, full_page=True)
                    except Exception:
                        pass
                    browser, context, page = recover_worker_page(
                        playwright,
                        browser=browser,
                        context=context,
                        page=page,
                        show_browser=show_browser,
                        block_assets=block_assets,
                        worker_id=worker_id,
                    )

            if completed:
                continue

            scenario_summaries.append(
                {
                    "row_num": scenario.row_num,
                    "result_url": last_result_url,
                    "usage": last_usage,
                    "plan_count": 0,
                    "status": "failed",
                    "error": last_error,
                }
            )

            error_note = f"Scenario failed after retries: {last_error.splitlines()[0][:150]}"
            failed_placeholder = build_no_plans_placeholder_row(
                scenario=scenario,
                execution_date=execution_date,
                usage=last_usage,
                result_url=last_result_url,
                note=error_note,
            )
            failed_placeholder["__row_num"] = str(scenario.row_num)
            failed_placeholder["__plan_order"] = "0"
            all_rows.append(failed_placeholder)

            if fail_fast:
                raise RuntimeError(
                    f"Worker {worker_id} failed on row {scenario.row_num} after {attempt_limit} attempts: {last_error}"
                )

        safe_close(page)
        safe_close(context)
        safe_close(browser)

    return all_rows, scenario_summaries


def split_scenarios_for_workers(scenarios: list[Scenario], workers: int) -> list[list[Scenario]]:
    chunks: list[list[Scenario]] = [[] for _ in range(max(workers, 1))]
    for idx, scenario in enumerate(scenarios):
        chunks[idx % max(workers, 1)].append(scenario)
    return [chunk for chunk in chunks if chunk]


def write_output_xlsx(output_file: Path, rows: list[dict[str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    ws.append(OUTPUT_COLUMNS)
    for row in rows:
        ws.append([row.get(col, "") for col in OUTPUT_COLUMNS])

    wb.save(output_file)


def write_output_csv(output_file: Path, rows: list[dict[str, str]]) -> None:
    import csv

    csv_file = output_file.with_suffix(".csv")
    with csv_file.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=OUTPUT_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def append_rows_to_csv(csv_file: Path, rows: list[dict[str, str]]) -> None:
    import csv

    def normalize_header_token(value: str) -> str:
        # Normalize header labels across legacy variants (case/spacing/BOM differences).
        return re.sub(r"[^a-z0-9]", "", clean_string(value).lower())

    csv_file.parent.mkdir(parents=True, exist_ok=True)

    file_exists = csv_file.exists()
    has_content = file_exists and csv_file.stat().st_size > 0
    append_mode = "dict"

    if has_content:
        with csv_file.open("r", encoding="utf-8", newline="") as f:
            reader = csv.reader(f)
            header = next(reader, [])

        normalized_existing = [normalize_header_token(col) for col in header if normalize_header_token(col)]
        normalized_expected = [normalize_header_token(col) for col in OUTPUT_COLUMNS]

        is_exact_logical_match = normalized_existing == normalized_expected
        is_legacy_positional_match = (
            len(header) == len(OUTPUT_COLUMNS)
            and normalized_existing == normalized_expected[: len(normalized_existing)]
        )

        if not is_exact_logical_match and not is_legacy_positional_match:
            raise ValueError(
                f"Cumulative CSV has unexpected header in {csv_file}. "
                "Expected exact PowerSwitch output columns."
            )

        if is_legacy_positional_match:
            append_mode = "positional"

    if not has_content:
        with csv_file.open("w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=OUTPUT_COLUMNS, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(rows)
        return

    with csv_file.open("a", encoding="utf-8", newline="") as f:
        if append_mode == "positional":
            writer = csv.writer(f)
            for row in rows:
                writer.writerow([row.get(col, "") for col in OUTPUT_COLUMNS])
            return

        writer = csv.DictWriter(f, fieldnames=OUTPUT_COLUMNS, extrasaction="ignore")
        writer.writerows(rows)


def parse_args() -> argparse.Namespace:
    today = datetime.now().strftime("%Y%m%d")
    parser = argparse.ArgumentParser(description="Run monthly PowerSwitch scrape from spec rows.")
    parser.add_argument(
        "--spec-file",
        default="20260720 Powerswitch Specs.xlsx",
        help="Path to input spec workbook containing columns A-G.",
    )
    parser.add_argument(
        "--output-file",
        default=f"{today} Powerswitch Output.xlsx",
        help="Output workbook path. If only a filename is provided, it is written to the Output folder.",
    )
    parser.add_argument(
        "--all-output-csv",
        default=str(DEFAULT_ALL_OUTPUT_CSV.name),
        help="Cumulative CSV to append on every run. If only a filename is provided, it is written to the Output folder.",
    )
    parser.add_argument(
        "--rows",
        type=int,
        default=5,
        help="Number of spec rows to scrape starting from --start-row.",
    )
    parser.add_argument(
        "--workers",
        type=int,
        default=1,
        help="Parallel worker count (default 1). Increase to reduce wall-clock runtime.",
    )
    parser.add_argument(
        "--start-row",
        type=int,
        default=2,
        help="Worksheet row number to start reading scenarios from.",
    )
    parser.add_argument(
        "--row-list",
        default="",
        help="Optional comma-separated row list/ranges to run (e.g. 101,105,200-205). Overrides --start-row/--rows.",
    )
    parser.add_argument(
        "--show-browser",
        action="store_true",
        help="Show browser window while scraping (headless by default).",
    )
    parser.add_argument(
        "--timeout-ms",
        type=int,
        default=45_000,
        help="Timeout in ms for most UI interactions.",
    )
    parser.add_argument(
        "--max-retries",
        type=int,
        default=2,
        help="Retry count per scenario before marking it as failed (default: 2).",
    )
    parser.add_argument(
        "--retry-empty-results",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="Retry once when a results page returns zero plans (default: enabled).",
    )
    parser.add_argument(
        "--fail-fast",
        action="store_true",
        help="Stop immediately when a scenario fails after retries.",
    )
    parser.add_argument(
        "--block-assets",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="Block image/font/media requests to speed up scraping (default: enabled).",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    spec_file = Path(args.spec_file)
    output_file = resolve_output_path(args.output_file)
    all_output_csv = resolve_output_path(args.all_output_csv)

    if not spec_file.exists():
        raise FileNotFoundError(f"Spec file not found: {spec_file}")
    if args.workers < 1:
        raise ValueError("--workers must be at least 1")
    if args.max_retries < 0:
        raise ValueError("--max-retries must be 0 or greater")

    output_file.parent.mkdir(parents=True, exist_ok=True)
    all_output_csv.parent.mkdir(parents=True, exist_ok=True)

    row_numbers: list[int] | None = None
    if args.row_list.strip():
        row_numbers = parse_row_list(args.row_list)

    scenarios = load_scenarios(
        spec_file,
        limit_rows=args.rows if row_numbers is None else len(row_numbers),
        start_row=args.start_row,
        row_numbers=row_numbers,
    )
    execution_date = datetime.now().strftime("%d/%m/%Y")

    all_rows: list[dict[str, str]] = []
    scenario_summaries: list[dict[str, str | int]] = []

    if args.workers == 1:
        all_rows, scenario_summaries = process_scenario_batch(
            scenarios=scenarios,
            timeout_ms=args.timeout_ms,
            show_browser=args.show_browser,
            execution_date=execution_date,
            block_assets=args.block_assets,
            worker_id=1,
            max_retries=args.max_retries,
            retry_empty_results=args.retry_empty_results,
            fail_fast=args.fail_fast,
        )
    else:
        chunks = split_scenarios_for_workers(scenarios, args.workers)
        print(f"Running {len(chunks)} parallel workers", flush=True)

        with ProcessPoolExecutor(max_workers=len(chunks)) as executor:
            futures = {
                executor.submit(
                    process_scenario_batch,
                    scenarios=chunk,
                    timeout_ms=args.timeout_ms,
                    show_browser=args.show_browser,
                    execution_date=execution_date,
                    block_assets=args.block_assets,
                    worker_id=idx + 1,
                    max_retries=args.max_retries,
                    retry_empty_results=args.retry_empty_results,
                    fail_fast=args.fail_fast,
                )
                : idx + 1
                for idx, chunk in enumerate(chunks)
            }

            for future in as_completed(futures):
                worker_id = futures[future]
                try:
                    rows, summaries = future.result()
                except Exception as exc:
                    if args.fail_fast:
                        raise
                    print(f"Worker {worker_id} failed: {type(exc).__name__}: {exc}", flush=True)
                    continue
                all_rows.extend(rows)
                scenario_summaries.extend(summaries)

    all_rows.sort(key=lambda row: (int(row.get("__row_num", "0")), int(row.get("__plan_order", "0"))))
    scenario_summaries.sort(key=lambda item: int(item["row_num"]))

    write_output_xlsx(output_file, all_rows)
    write_output_csv(output_file, all_rows)
    append_rows_to_csv(all_output_csv, all_rows)

    print("\nScrape complete", flush=True)
    print(f"Execution date: {execution_date}", flush=True)
    print(f"Scenarios run: {len(scenarios)}", flush=True)
    print(f"Workers used: {args.workers}", flush=True)
    print(f"Total output rows: {len(all_rows)}", flush=True)
    failed_summaries = [s for s in scenario_summaries if s.get("status") == "failed"]
    empty_summaries = [s for s in scenario_summaries if s.get("status") == "empty"]
    print(f"Failed scenarios: {len(failed_summaries)}", flush=True)
    print(f"Empty-result scenarios: {len(empty_summaries)}", flush=True)
    print(f"Output workbook: {output_file}", flush=True)
    print(f"Output CSV: {output_file.with_suffix('.csv')}", flush=True)
    print(f"Cumulative CSV (appended): {all_output_csv}", flush=True)
    print("Scenario summary:", flush=True)
    print(json.dumps(scenario_summaries, indent=2), flush=True)


if __name__ == "__main__":
    main()
