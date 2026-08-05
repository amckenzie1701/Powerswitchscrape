from __future__ import annotations

import argparse
import csv
import html
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from statistics import median
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

OUTPUT_COLUMNS = [
    "Execution Date",
    "Network Code",
    "Subnetwork",
    "Region",
    "Area",
    "People",
    "Water heating",
    "General heating",
    "Usage kWh",
    "Retailer",
    "Plan Name",
    "Tariffs and discounts",
    "Plan attributes",
    "Estimated yearly cost",
    "Prices last changed",
    "Result URL",
]

SCENARIO_COLUMNS = [
    "Network Code",
    "Subnetwork",
    "Region",
    "Area",
    "People",
    "Water heating",
    "General heating",
]

NO_PLANS_LABEL = "No Plans"


@dataclass
class DatasetSummary:
    total_rows: int
    offer_rows: int
    placeholder_rows: int
    unique_scenarios: int
    unique_retailers: int
    unique_regions: int
    unique_areas: int
    avg_cost: float | None
    median_cost: float | None
    min_cost: float | None
    max_cost: float | None


def clean_string(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_currency(value: str) -> float | None:
    text = clean_string(value)
    if not text:
        return None
    match = re.search(r"([0-9][0-9,]*(?:\.[0-9]+)?)", text.replace("$", ""))
    if not match:
        return None
    return float(match.group(1).replace(",", ""))


def scenario_key(row: dict[str, str]) -> tuple[str, ...]:
    return tuple(clean_string(row.get(col, "")) for col in SCENARIO_COLUMNS)


def is_offer_row(row: dict[str, str]) -> bool:
    plan_name = clean_string(row.get("Plan Name", ""))
    return bool(plan_name) and plan_name != NO_PLANS_LABEL


def is_placeholder_row(row: dict[str, str]) -> bool:
    return clean_string(row.get("Plan Name", "")) == NO_PLANS_LABEL


def load_rows(input_file: Path) -> list[dict[str, str]]:
    suffix = input_file.suffix.lower()
    if suffix == ".csv":
        with input_file.open("r", encoding="utf-8", newline="") as f:
            reader = csv.DictReader(f)
            rows: list[dict[str, str]] = []
            for row in reader:
                rows.append({col: clean_string(row.get(col, "")) for col in OUTPUT_COLUMNS})
            return rows

    if suffix not in {".xlsx", ".xlsm"}:
        raise ValueError(f"Unsupported input format: {input_file}")

    # Some workbooks have stale dimension metadata; read_only mode can truncate rows.
    wb = load_workbook(input_file, data_only=True)
    ws = wb.active
    if ws.max_row < 1:
        return []

    headers = [clean_string(ws.cell(row=1, column=col_idx).value) for col_idx in range(1, ws.max_column + 1)]
    col_index = {header: idx for idx, header in enumerate(headers)}
    missing = [col for col in OUTPUT_COLUMNS if col not in col_index]
    if missing:
        raise ValueError(f"File {input_file} is missing required columns: {missing}")

    rows: list[dict[str, str]] = []
    for row_idx in range(2, ws.max_row + 1):
        row_dict: dict[str, str] = {}
        for col in OUTPUT_COLUMNS:
            col_idx = col_index[col] + 1
            value = ws.cell(row=row_idx, column=col_idx).value
            row_dict[col] = clean_string(value)
        if not any(row_dict.values()):
            continue
        rows.append(row_dict)

    return rows


def average(values: list[float]) -> float | None:
    if not values:
        return None
    return sum(values) / len(values)


def build_dataset_summary(rows: list[dict[str, str]]) -> DatasetSummary:
    offers = [row for row in rows if is_offer_row(row)]
    placeholders = [row for row in rows if is_placeholder_row(row)]

    costs = [cost for cost in (parse_currency(row.get("Estimated yearly cost", "")) for row in offers) if cost is not None]

    retailers = {
        clean_string(row.get("Retailer", ""))
        for row in offers
        if clean_string(row.get("Retailer", ""))
    }
    regions = {clean_string(row.get("Region", "")) for row in rows if clean_string(row.get("Region", ""))}
    areas = {clean_string(row.get("Area", "")) for row in rows if clean_string(row.get("Area", ""))}
    scenarios = {scenario_key(row) for row in rows}

    return DatasetSummary(
        total_rows=len(rows),
        offer_rows=len(offers),
        placeholder_rows=len(placeholders),
        unique_scenarios=len(scenarios),
        unique_retailers=len(retailers),
        unique_regions=len(regions),
        unique_areas=len(areas),
        avg_cost=average(costs),
        median_cost=median(costs) if costs else None,
        min_cost=min(costs) if costs else None,
        max_cost=max(costs) if costs else None,
    )


def metric_delta(current: float | int | None, baseline: float | int | None) -> float | int | None:
    if current is None or baseline is None:
        return None
    return current - baseline


def metric_delta_pct(current: float | int | None, baseline: float | int | None) -> float | None:
    if current is None or baseline in (None, 0):
        return None
    return (float(current) - float(baseline)) / float(baseline)


def build_overview_rows(current_rows: list[dict[str, str]], baseline_rows: list[dict[str, str]]) -> list[list[Any]]:
    current = build_dataset_summary(current_rows)
    baseline = build_dataset_summary(baseline_rows)

    metrics: list[tuple[str, float | int | None, float | int | None]] = [
        ("Total output rows", current.total_rows, baseline.total_rows),
        ("Offer rows (excluding No Plans)", current.offer_rows, baseline.offer_rows),
        ("No Plans placeholder rows", current.placeholder_rows, baseline.placeholder_rows),
        ("Unique scenarios", current.unique_scenarios, baseline.unique_scenarios),
        ("Unique retailers", current.unique_retailers, baseline.unique_retailers),
        ("Unique regions", current.unique_regions, baseline.unique_regions),
        ("Unique areas", current.unique_areas, baseline.unique_areas),
        ("Average estimated yearly cost", current.avg_cost, baseline.avg_cost),
        ("Median estimated yearly cost", current.median_cost, baseline.median_cost),
        ("Minimum estimated yearly cost", current.min_cost, baseline.min_cost),
        ("Maximum estimated yearly cost", current.max_cost, baseline.max_cost),
    ]

    rows: list[list[Any]] = []
    for metric, current_value, baseline_value in metrics:
        rows.append(
            [
                metric,
                current_value,
                baseline_value,
                metric_delta(current_value, baseline_value),
                metric_delta_pct(current_value, baseline_value),
            ]
        )
    return rows


def build_group_stats(rows: list[dict[str, str]], group_cols: list[str]) -> dict[tuple[str, ...], dict[str, Any]]:
    grouped: dict[tuple[str, ...], dict[str, Any]] = defaultdict(
        lambda: {
            "scenarios": set(),
            "offer_rows": 0,
            "no_plan_scenarios": set(),
            "costs": [],
        }
    )

    for row in rows:
        key = tuple(clean_string(row.get(col, "")) for col in group_cols)
        skey = scenario_key(row)
        item = grouped[key]
        item["scenarios"].add(skey)

        if is_offer_row(row):
            item["offer_rows"] += 1
            cost = parse_currency(row.get("Estimated yearly cost", ""))
            if cost is not None:
                item["costs"].append(cost)
        elif is_placeholder_row(row):
            item["no_plan_scenarios"].add(skey)

    summary: dict[tuple[str, ...], dict[str, Any]] = {}
    for key, item in grouped.items():
        costs = item["costs"]
        summary[key] = {
            "scenarios": len(item["scenarios"]),
            "offer_rows": item["offer_rows"],
            "no_plan_scenarios": len(item["no_plan_scenarios"]),
            "avg_cost": average(costs),
            "median_cost": median(costs) if costs else None,
        }
    return summary


def build_group_comparison_rows(
    current_rows: list[dict[str, str]],
    baseline_rows: list[dict[str, str]],
    group_cols: list[str],
) -> tuple[list[str], list[list[Any]]]:
    current_stats = build_group_stats(current_rows, group_cols)
    baseline_stats = build_group_stats(baseline_rows, group_cols)

    keys = sorted(
        set(current_stats) | set(baseline_stats),
        key=lambda key: tuple(part.lower() for part in key),
    )

    headers = [
        *group_cols,
        "Current Scenarios",
        "Baseline Scenarios",
        "Scenario Delta",
        "Current Offer Rows",
        "Baseline Offer Rows",
        "Offer Row Delta",
        "Current No Plans Scenarios",
        "Baseline No Plans Scenarios",
        "No Plans Scenario Delta",
        "Current Avg Cost",
        "Baseline Avg Cost",
        "Avg Cost Delta",
        "Current Median Cost",
        "Baseline Median Cost",
        "Median Cost Delta",
    ]

    rows: list[list[Any]] = []
    for key in keys:
        current = current_stats.get(
            key,
            {
                "scenarios": 0,
                "offer_rows": 0,
                "no_plan_scenarios": 0,
                "avg_cost": None,
                "median_cost": None,
            },
        )
        baseline = baseline_stats.get(
            key,
            {
                "scenarios": 0,
                "offer_rows": 0,
                "no_plan_scenarios": 0,
                "avg_cost": None,
                "median_cost": None,
            },
        )

        rows.append(
            [
                *key,
                current["scenarios"],
                baseline["scenarios"],
                current["scenarios"] - baseline["scenarios"],
                current["offer_rows"],
                baseline["offer_rows"],
                current["offer_rows"] - baseline["offer_rows"],
                current["no_plan_scenarios"],
                baseline["no_plan_scenarios"],
                current["no_plan_scenarios"] - baseline["no_plan_scenarios"],
                current["avg_cost"],
                baseline["avg_cost"],
                metric_delta(current["avg_cost"], baseline["avg_cost"]),
                current["median_cost"],
                baseline["median_cost"],
                metric_delta(current["median_cost"], baseline["median_cost"]),
            ]
        )

    rows.sort(key=lambda row: (-int(row[len(group_cols) + 3]), *(str(v).lower() for v in row[: len(group_cols)])))
    return headers, rows


def build_retailer_rankings(current_rows: list[dict[str, str]], baseline_rows: list[dict[str, str]]) -> tuple[list[str], list[list[Any]]]:
    def retailer_stats(rows: list[dict[str, str]]) -> dict[str, dict[str, Any]]:
        stats: dict[str, dict[str, Any]] = defaultdict(lambda: {"offers": 0, "costs": []})
        for row in rows:
            if not is_offer_row(row):
                continue
            retailer = clean_string(row.get("Retailer", ""))
            if not retailer:
                retailer = "(Unknown)"
            stats[retailer]["offers"] += 1
            cost = parse_currency(row.get("Estimated yearly cost", ""))
            if cost is not None:
                stats[retailer]["costs"].append(cost)

        out: dict[str, dict[str, Any]] = {}
        for retailer, values in stats.items():
            out[retailer] = {
                "offers": values["offers"],
                "avg_cost": average(values["costs"]),
            }
        return out

    def dense_rank(offers_by_name: dict[str, dict[str, Any]]) -> dict[str, int]:
        ordered = sorted(offers_by_name.items(), key=lambda item: (-int(item[1]["offers"]), item[0].lower()))
        ranks: dict[str, int] = {}
        current_rank = 0
        previous_offers: int | None = None
        for idx, (name, values) in enumerate(ordered, start=1):
            offers = int(values["offers"])
            if previous_offers is None or offers != previous_offers:
                current_rank = idx
                previous_offers = offers
            ranks[name] = current_rank
        return ranks

    current = retailer_stats(current_rows)
    baseline = retailer_stats(baseline_rows)
    current_rank = dense_rank(current)
    baseline_rank = dense_rank(baseline)

    all_retailers = sorted(
        set(current) | set(baseline),
        key=lambda name: (-int(current.get(name, {}).get("offers", 0)), name.lower()),
    )

    headers = [
        "Current Rank",
        "Baseline Rank",
        "Rank Change",
        "Retailer",
        "Current Offer Rows",
        "Baseline Offer Rows",
        "Offer Row Delta",
        "Current Avg Cost",
        "Baseline Avg Cost",
        "Avg Cost Delta",
    ]

    rows: list[list[Any]] = []
    for retailer in all_retailers:
        current_values = current.get(retailer, {"offers": 0, "avg_cost": None})
        baseline_values = baseline.get(retailer, {"offers": 0, "avg_cost": None})
        c_rank = current_rank.get(retailer)
        b_rank = baseline_rank.get(retailer)

        rank_change: int | None = None
        if c_rank is not None and b_rank is not None:
            rank_change = b_rank - c_rank

        rows.append(
            [
                c_rank,
                b_rank,
                rank_change,
                retailer,
                current_values["offers"],
                baseline_values["offers"],
                int(current_values["offers"]) - int(baseline_values["offers"]),
                current_values["avg_cost"],
                baseline_values["avg_cost"],
                metric_delta(current_values["avg_cost"], baseline_values["avg_cost"]),
            ]
        )

    return headers, rows


def build_scenario_comparison_rows(
    current_rows: list[dict[str, str]],
    baseline_rows: list[dict[str, str]],
) -> tuple[list[str], list[list[Any]]]:
    def scenario_stats(rows: list[dict[str, str]]) -> dict[tuple[str, ...], dict[str, Any]]:
        stats: dict[tuple[str, ...], dict[str, Any]] = defaultdict(
            lambda: {
                "offer_rows": 0,
                "has_no_plans": False,
                "costs": [],
            }
        )

        for row in rows:
            key = scenario_key(row)
            item = stats[key]
            if is_offer_row(row):
                item["offer_rows"] += 1
                cost = parse_currency(row.get("Estimated yearly cost", ""))
                if cost is not None:
                    item["costs"].append(cost)
            elif is_placeholder_row(row):
                item["has_no_plans"] = True

        out: dict[tuple[str, ...], dict[str, Any]] = {}
        for key, values in stats.items():
            out[key] = {
                "offer_rows": values["offer_rows"],
                "has_no_plans": values["has_no_plans"],
                "avg_cost": average(values["costs"]),
            }
        return out

    current = scenario_stats(current_rows)
    baseline = scenario_stats(baseline_rows)
    keys = sorted(set(current) | set(baseline), key=lambda key: tuple(part.lower() for part in key))

    headers = [
        *SCENARIO_COLUMNS,
        "Current Offer Rows",
        "Baseline Offer Rows",
        "Offer Row Delta",
        "Current Has No Plans",
        "Baseline Has No Plans",
        "Current Avg Cost",
        "Baseline Avg Cost",
        "Avg Cost Delta",
    ]

    rows: list[list[Any]] = []
    for key in keys:
        c = current.get(key, {"offer_rows": 0, "has_no_plans": False, "avg_cost": None})
        b = baseline.get(key, {"offer_rows": 0, "has_no_plans": False, "avg_cost": None})
        rows.append(
            [
                *key,
                c["offer_rows"],
                b["offer_rows"],
                int(c["offer_rows"]) - int(b["offer_rows"]),
                "Yes" if c["has_no_plans"] else "No",
                "Yes" if b["has_no_plans"] else "No",
                c["avg_cost"],
                b["avg_cost"],
                metric_delta(c["avg_cost"], b["avg_cost"]),
            ]
        )

    rows.sort(key=lambda row: (-int(row[len(SCENARIO_COLUMNS)]), *(str(v).lower() for v in row[: len(SCENARIO_COLUMNS)])))
    return headers, rows


def add_sheet(
    wb: Workbook,
    title: str,
    headers: list[str],
    rows: list[list[Any]],
    currency_columns: set[str] | None = None,
    percent_columns: set[str] | None = None,
) -> None:
    ws = wb.create_sheet(title=title)
    ws.append(headers)
    for row in rows:
        ws.append(row)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx).font = Font(bold=True)

    currency_columns = currency_columns or set()
    percent_columns = percent_columns or set()

    for col_idx, header in enumerate(headers, start=1):
        if header in currency_columns:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = "$#,##0.00"
        if header in percent_columns:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = "0.00%"


def format_value_for_display(
        value: Any,
        header: str,
        currency_columns: set[str],
        percent_columns: set[str],
) -> str:
        if value is None:
                return ""

        if header in percent_columns and isinstance(value, (int, float)):
                return f"{float(value):.2%}"

        if header in currency_columns and isinstance(value, (int, float)):
                return f"${float(value):,.2f}"

        if isinstance(value, int):
                return f"{value:,}"

        if isinstance(value, float):
                if value.is_integer():
                        return f"{int(value):,}"
                return f"{value:,.2f}"

        return str(value)


def write_html_report(
        output_file: Path,
        sections: list[dict[str, Any]],
        current_file: Path,
        baseline_file: Path,
) -> None:
        generated = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        parts: list[str] = []
        parts.append("<!DOCTYPE html>")
        parts.append("<html lang=\"en\">")
        parts.append("<head>")
        parts.append("<meta charset=\"utf-8\" />")
        parts.append("<meta name=\"viewport\" content=\"width=device-width, initial-scale=1\" />")
        parts.append("<title>PowerSwitch Comparison Report</title>")
        parts.append(
                """
<style>
    :root {
        --bg: #f3f5f7;
        --card: #ffffff;
        --ink: #13222f;
        --muted: #5b6b79;
        --line: #d8e0e7;
        --accent: #0b7285;
        --accent-soft: #e8f5f8;
    }
    body {
        margin: 0;
        font-family: "Segoe UI", "Helvetica Neue", sans-serif;
        background: linear-gradient(180deg, #ecf2f6 0%, #f8fafb 100%);
        color: var(--ink);
    }
    .wrap {
        max-width: 1500px;
        margin: 0 auto;
        padding: 20px;
    }
    .hero {
        background: var(--card);
        border: 1px solid var(--line);
        border-radius: 14px;
        padding: 18px 20px;
        margin-bottom: 18px;
    }
    .hero h1 {
        margin: 0 0 8px 0;
        font-size: 28px;
    }
    .meta {
        display: grid;
        grid-template-columns: repeat(auto-fit, minmax(240px, 1fr));
        gap: 8px;
        color: var(--muted);
        font-size: 14px;
    }
    .section {
        background: var(--card);
        border: 1px solid var(--line);
        border-radius: 14px;
        padding: 14px;
        margin-bottom: 14px;
    }
    .section h2 {
        margin: 0;
        font-size: 18px;
    }
    .controls {
        margin: 10px 0;
    }
    .controls input {
        width: min(460px, 100%);
        border: 1px solid var(--line);
        border-radius: 8px;
        padding: 8px 10px;
        font-size: 14px;
    }
    .table-wrap {
        overflow: auto;
        max-height: 62vh;
        border: 1px solid var(--line);
        border-radius: 10px;
        background: #fff;
    }
    table {
        border-collapse: collapse;
        width: 100%;
        min-width: 920px;
    }
    th, td {
        padding: 7px 10px;
        border-bottom: 1px solid #edf1f5;
        text-align: left;
        font-size: 13px;
        white-space: nowrap;
    }
    th {
        position: sticky;
        top: 0;
        z-index: 1;
        background: var(--accent-soft);
        border-bottom: 1px solid #c8dce2;
        font-weight: 700;
    }
    tbody tr:nth-child(even) {
        background: #fbfdff;
    }
    .note {
        color: var(--muted);
        font-size: 12px;
        margin-top: 8px;
    }
</style>
                """
        )
        parts.append("</head>")
        parts.append("<body>")
        parts.append("<div class=\"wrap\">")
        parts.append("<section class=\"hero\">")
        parts.append("<h1>PowerSwitch Comparison Report</h1>")
        parts.append("<div class=\"meta\">")
        parts.append(f"<div><strong>Generated:</strong> {html.escape(generated)}</div>")
        parts.append(f"<div><strong>Current File:</strong> {html.escape(str(current_file))}</div>")
        parts.append(f"<div><strong>Baseline File:</strong> {html.escape(str(baseline_file))}</div>")
        parts.append("</div>")
        parts.append("</section>")

        for idx, section in enumerate(sections, start=1):
                title = section["title"]
                headers = section["headers"]
                rows = section["rows"]
                currency_columns = set(section.get("currency_columns") or [])
                percent_columns = set(section.get("percent_columns") or [])
                table_id = f"table_{idx}"
                search_id = f"search_{idx}"

                parts.append("<section class=\"section\">")
                parts.append(f"<h2>{html.escape(title)}</h2>")
                parts.append("<div class=\"controls\">")
                parts.append(
                        f"<input id=\"{search_id}\" type=\"text\" placeholder=\"Filter this table...\" oninput=\"filterTable('{table_id}', this.value)\" />"
                )
                parts.append("</div>")
                parts.append("<div class=\"table-wrap\">")
                parts.append(f"<table id=\"{table_id}\">")
                parts.append("<thead><tr>")
                for header in headers:
                        parts.append(f"<th>{html.escape(str(header))}</th>")
                parts.append("</tr></thead>")
                parts.append("<tbody>")
                for row in rows:
                        parts.append("<tr>")
                        for col_idx, value in enumerate(row):
                                header = headers[col_idx]
                                display = format_value_for_display(value, header, currency_columns, percent_columns)
                                parts.append(f"<td>{html.escape(display)}</td>")
                        parts.append("</tr>")
                parts.append("</tbody>")
                parts.append("</table>")
                parts.append("</div>")
                parts.append(f"<div class=\"note\">Rows: {len(rows):,}</div>")
                parts.append("</section>")

        parts.append(
                """
<script>
    function filterTable(tableId, query) {
        const q = (query || '').toLowerCase();
        const table = document.getElementById(tableId);
        if (!table) return;
        const rows = table.tBodies[0].rows;
        for (const row of rows) {
            const text = row.innerText.toLowerCase();
            row.style.display = text.includes(q) ? '' : 'none';
        }
    }
</script>
                """
        )
        parts.append("</div>")
        parts.append("</body>")
        parts.append("</html>")

        output_file.write_text("\n".join(parts), encoding="utf-8")


def parse_args() -> argparse.Namespace:
    today = datetime.now().strftime("%Y%m%d")
    parser = argparse.ArgumentParser(
        description="Build top-level comparison tables between a current and baseline PowerSwitch output file."
    )
    parser.add_argument(
        "--current-file",
        default="Output/ALLPOWERSWITCHOUT.csv",
        help="Current scrape file (.xlsx or .csv). Defaults to cumulative Output/ALLPOWERSWITCHOUT.csv.",
    )
    parser.add_argument(
        "--baseline-file",
        default="Expected Output.xlsx",
        help="Baseline file to compare against (.xlsx or .csv).",
    )
    parser.add_argument(
        "--output-file",
        default=f"{today} Powerswitch Comparison Tables.xlsx",
        help="Output workbook for comparison tables.",
    )
    parser.add_argument(
        "--html-file",
        default=f"{today} Powerswitch Comparison Tables.html",
        help="Output HTML report for comparison tables.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    current_file = Path(args.current_file)
    baseline_file = Path(args.baseline_file)
    output_file = Path(args.output_file)
    html_file = Path(args.html_file)

    if not current_file.exists():
        raise FileNotFoundError(f"Current file not found: {current_file}")
    if not baseline_file.exists():
        raise FileNotFoundError(f"Baseline file not found: {baseline_file}")

    current_rows = load_rows(current_file)
    baseline_rows = load_rows(baseline_file)

    sections: list[dict[str, Any]] = []

    sections.append(
        {
            "title": "Overview",
            "headers": ["Metric", "Current", "Baseline", "Delta", "Delta %"],
            "rows": build_overview_rows(current_rows, baseline_rows),
            "currency_columns": set(),
            "percent_columns": {"Delta %"},
        }
    )

    retailer_headers, retailer_rows = build_retailer_rankings(current_rows, baseline_rows)
    sections.append(
        {
            "title": "Rankings_Retailer",
            "headers": retailer_headers,
            "rows": retailer_rows,
            "currency_columns": {"Current Avg Cost", "Baseline Avg Cost", "Avg Cost Delta"},
            "percent_columns": set(),
        }
    )

    for title, cols in [
        ("Compare_Region", ["Region"]),
        ("Compare_Area", ["Area"]),
        ("Compare_People", ["People"]),
        ("Compare_Water", ["Water heating"]),
        ("Compare_Heating", ["General heating"]),
    ]:
        headers, rows = build_group_comparison_rows(current_rows, baseline_rows, cols)
        sections.append(
            {
                "title": title,
                "headers": headers,
                "rows": rows,
                "currency_columns": {
                    "Current Avg Cost",
                    "Baseline Avg Cost",
                    "Avg Cost Delta",
                    "Current Median Cost",
                    "Baseline Median Cost",
                    "Median Cost Delta",
                },
                "percent_columns": set(),
            }
        )

    scenario_headers, scenario_rows = build_scenario_comparison_rows(current_rows, baseline_rows)
    sections.append(
        {
            "title": "Scenario_Compare",
            "headers": scenario_headers,
            "rows": scenario_rows,
            "currency_columns": {"Current Avg Cost", "Baseline Avg Cost", "Avg Cost Delta"},
            "percent_columns": set(),
        }
    )

    wb = Workbook()
    wb.remove(wb.active)
    for section in sections:
        add_sheet(
            wb,
            section["title"],
            section["headers"],
            section["rows"],
            currency_columns=set(section.get("currency_columns") or []),
            percent_columns=set(section.get("percent_columns") or []),
        )

    wb.save(output_file)
    write_html_report(
        output_file=html_file,
        sections=sections,
        current_file=current_file,
        baseline_file=baseline_file,
    )

    print("Comparison workbook created")
    print(f"Current file: {current_file}")
    print(f"Baseline file: {baseline_file}")
    print(f"Rows in current: {len(current_rows)}")
    print(f"Rows in baseline: {len(baseline_rows)}")
    print(f"Output workbook: {output_file}")
    print(f"Output HTML: {html_file}")


if __name__ == "__main__":
    main()
